#!/usr/bin/env python3
"""
Generate two demo Excel files for the deterministic demo scenario:

1. market_research_trends_2025.xlsx  — Trend analysis, competitor pricing, consumer sentiment
2. inventory_dead_stock_report.xlsx  — Dead stock summary, Tuscany Collection detail, trend matching

These files are uploaded to Google Drive and referenced in the agent's canned responses.
"""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("ERROR: openpyxl is required.\n  pip install openpyxl")

OUTPUT_DIR = Path(__file__).parent

# ── Shared styles ──────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Calibri", size=10)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="B4C6E7"))
ACCENT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow highlight
RED_FONT = Font(name="Calibri", size=10, color="CC0000", bold=True)
GREEN_FONT = Font(name="Calibri", size=10, color="006600", bold=True)


def _style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _style_data(ws, num_rows, num_cols):
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL


def _autofit(ws, widths: dict):
    for col_idx, (_, width) in enumerate(widths.items(), start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# =======================================================================
# FILE 1: market_research_trends_2025.xlsx
# =======================================================================
def create_market_research():
    wb = Workbook()

    # ── Sheet 1: Trend Analysis ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Trend Analysis"
    headers1 = [
        "Trend Keyword", "Category", "Google Trends Index (0-100)",
        "YoY Growth %", "Peak Month", "Trend Status", "Related Searches",
    ]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    _style_header(ws1, len(headers1))

    trends_data = [
        ["Organic Modern", "Interior Design", 92, 187, "Jan 2025", "🔥 SPIKE", "organic modern living room, organic modern furniture"],
        ["Raw Wood Furniture", "Furniture", 84, 145, "Dec 2024", "🔥 SPIKE", "raw oak table, raw wood dining, unfinished wood"],
        ["Boucle Fabric", "Textiles", 78, 112, "Nov 2024", "📈 Rising", "boucle chair, boucle sofa, textured upholstery"],
        ["Warm Neutrals", "Color Palette", 88, 156, "Jan 2025", "🔥 SPIKE", "warm neutral living room, cream beige decor"],
        ["Japandi Style", "Interior Design", 65, 42, "Oct 2024", "📊 Stable", "japandi furniture, minimalist warm"],
        ["Cottagecore", "Interior Design", 45, -23, "Jun 2024", "📉 Declining", "cottagecore decor, rustic farmhouse"],
        ["Mid-Century Modern", "Interior Design", 58, -8, "Sep 2024", "📊 Stable", "mcm furniture, retro modern"],
        ["Natural Linen", "Textiles", 76, 98, "Jan 2025", "📈 Rising", "linen upholstery, natural linen chair"],
        ["Sustainable Furniture", "Furniture", 71, 89, "Dec 2024", "📈 Rising", "eco furniture, sustainable wood"],
        ["Biophilic Design", "Interior Design", 69, 67, "Nov 2024", "📈 Rising", "nature-inspired interiors, indoor plants"],
    ]
    for r, row_data in enumerate(trends_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c == 3 and val >= 80:
                cell.font = GREEN_FONT
            if c == 6 and "SPIKE" in str(val):
                cell.fill = ACCENT_FILL

    _style_data(ws1, len(trends_data), len(headers1))
    _autofit(ws1, {
        "Trend Keyword": 22, "Category": 18, "Google Trends Index": 24,
        "YoY Growth %": 14, "Peak Month": 14, "Trend Status": 16, "Related Searches": 50,
    })
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{len(trends_data)+1}"

    # ── Sheet 2: Competitor Pricing ────────────────────────────────────
    ws2 = wb.create_sheet("Competitor Pricing")
    headers2 = [
        "Competitor", "Product Name", "Category", "Material",
        "Price (USD)", "Rating", "Reviews", "In Stock", "URL",
    ]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, len(headers2))

    competitor_data = [
        ["Competitor A", "Solid Wood Dining Table", "Dining Table", "Raw Oak", 1299, 4.5, 342, "Yes", "https://example.com/product-a"],
        ["Competitor B", "Raw Oak Dining Table", "Dining Table", "Raw Oak", 1199, 4.3, 218, "Yes", "https://example.com/product-b"],
        ["Competitor C", "Natural Dining Table", "Dining Table", "Raw Oak / Steel", 1499, 4.7, 567, "Yes", "https://example.com/product-c"],
        ["Competitor D", "Dining Table D", "Dining Table", "Raw Oak", 1349, 4.4, 423, "Yes", "https://example.com/product-d"],
        ["Competitor E", "Dining Table E", "Dining Table", "Raw Oak", 1149, 4.6, 189, "Yes", "https://example.com/product-e"],
        ["Competitor F", "Dining Table F", "Dining Table", "Oak Veneer", 549, 4.2, 1204, "Yes", "https://example.com/product-f"],
        ["Competitor G", "Salvaged Wood Trestle Table", "Dining Table", "Salvaged Oak", 2495, 4.8, 156, "Low Stock", "https://example.com/product-g"],
        ["Competitor A", "Linen Dining Chair", "Chair", "Linen / Oak", 449, 4.3, 287, "Yes", "https://example.com/chair-a"],
        ["Competitor B", "Woven Linen Dining Chair", "Chair", "Linen", 399, 4.1, 145, "Yes", "https://example.com/chair-b"],
        ["Competitor E", "Linen Chair E", "Chair", "Linen / Walnut", 379, 4.5, 203, "Yes", "https://example.com/chair-e"],
    ]
    for r, row_data in enumerate(competitor_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c == 5:
                cell.number_format = '"$"#,##0'
                if val >= 1200:
                    cell.font = RED_FONT

    _style_data(ws2, len(competitor_data), len(headers2))
    _autofit(ws2, {
        "Competitor": 22, "Product Name": 36, "Category": 16, "Material": 18,
        "Price (USD)": 14, "Rating": 10, "Reviews": 10, "In Stock": 12, "URL": 40,
    })
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(competitor_data)+1}"

    # ── Sheet 3: Consumer Sentiment ────────────────────────────────────
    ws3 = wb.create_sheet("Consumer Sentiment")
    headers3 = [
        "Platform", "Search Term / Hashtag", "Monthly Volume",
        "Engagement Rate %", "Sentiment (Positive %)", "Trend Direction", "Notes",
    ]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    _style_header(ws3, len(headers3))

    sentiment_data = [
        ["Pinterest", "#OrganicModern", 2400000, 8.7, 94, "📈 +67% MoM", "Top 5 trending home category"],
        ["Pinterest", "raw wood dining table", 890000, 6.2, 91, "📈 +45% MoM", "Saves up 3x vs last quarter"],
        ["Instagram", "#organicmodernhome", 1100000, 5.4, 89, "📈 +52% MoM", "Creator content surging"],
        ["Instagram", "#rawwood", 670000, 4.1, 87, "📈 +38% MoM", "Strong in 25-40 demographic"],
        ["TikTok", "organic modern room tour", 3200000, 12.3, 92, "🔥 Viral", "Multiple 1M+ view videos"],
        ["TikTok", "raw oak table makeover", 1800000, 9.8, 95, "🔥 Viral", "DIY + purchase intent high"],
        ["Google", "organic modern furniture buy", 340000, "N/A", "N/A", "📈 +187% YoY", "High commercial intent"],
        ["Google", "raw oak dining table near me", 220000, "N/A", "N/A", "📈 +145% YoY", "Local purchase intent"],
        ["YouTube", "organic modern home tour 2025", 890000, 7.6, 88, "📈 +89% YoY", "Long-form content growing"],
        ["Reddit", "r/interiordesign organic modern", 45000, 15.2, 82, "📈 +34% MoM", "Highly engaged community"],
    ]
    for r, row_data in enumerate(sentiment_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            if c == 3 and isinstance(val, (int, float)) and val >= 1000000:
                cell.font = GREEN_FONT

    _style_data(ws3, len(sentiment_data), len(headers3))
    _autofit(ws3, {
        "Platform": 14, "Search Term / Hashtag": 34, "Monthly Volume": 16,
        "Engagement Rate %": 18, "Sentiment (Positive %)": 20, "Trend Direction": 16, "Notes": 36,
    })
    ws3.freeze_panes = "A2"

    path = OUTPUT_DIR / "market_research_trends_2025.xlsx"
    wb.save(path)
    print(f"  ✓ {path}")
    return path


# =======================================================================
# FILE 2: inventory_dead_stock_report.xlsx
# =======================================================================
def create_inventory_report():
    wb = Workbook()

    # ── Sheet 1: Dead Stock Summary ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Dead Stock Summary"
    headers1 = [
        "SKU", "Product Name", "Collection", "Category", "Material",
        "Warehouse", "Qty On Hand", "Days Since Last Sale",
        "Stock Health Tier", "Clearance Price", "Unit Cost",
        "Capital Tied Up", "Recommendation",
    ]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    _style_header(ws1, len(headers1))

    dead_stock_data = [
        ["SKU-0001", "Tuscany Raw Oak Dining Table", "Tuscany Collection", "Dining Table", "raw oak",
         "WH-GA (Georgia)", 38, 210, "🔴 Dead Stock", 400, 320, 12160,
         "RELAUNCH — matches Organic Modern trend"],
        ["SKU-0002", "Tuscany Linen Dining Chair", "Tuscany Collection", "Chair", "linen",
         "WH-GA (Georgia)", 124, 210, "🔴 Dead Stock", 120, 95, 11780,
         "RELAUNCH — matches Organic Modern trend"],
        ["SKU-0038", "Metro Edge Steel Desk", "Metro Edge", "Desk", "steel",
         "WH-TX (Texas)", 15, 185, "🔴 Dead Stock", 280, 210, 3150,
         "Clearance — low trend alignment"],
        ["SKU-0041", "Coastal Breeze Rattan Bookshelf", "Coastal Breeze", "Bookshelf", "rattan",
         "WH-CA (California)", 22, 167, "🔴 Dead Stock", 190, 145, 3190,
         "Clearance — moderate trend alignment"],
        ["SKU-0045", "Nordic Frost Birch Console", "Nordic Frost", "Console", "birch",
         "WH-NY (New York)", 9, 152, "🔴 Dead Stock", 310, 240, 2160,
         "Clearance — low trend alignment"],
        ["SKU-0019", "Organic Living Bamboo Side Table", "Organic Living", "Side Table", "bamboo",
         "WH-IL (Illinois)", 31, 140, "🔴 Dead Stock", 95, 68, 2108,
         "Bundle with trending items"],
        ["SKU-0027", "Heritage Craft Walnut Nightstand", "Heritage Craft", "Nightstand", "walnut",
         "WH-GA (Georgia)", 18, 135, "🔴 Dead Stock", 175, 130, 2340,
         "Clearance — declining style"],
    ]
    for r, row_data in enumerate(dead_stock_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c == 10 or c == 11:
                cell.number_format = '"$"#,##0'
            if c == 12:
                cell.number_format = '"$"#,##0'
                cell.font = RED_FONT
            if c == 9:
                cell.font = RED_FONT
            # Highlight Tuscany rows
            if r in (2, 3):
                cell.fill = ACCENT_FILL

    _style_data(ws1, len(dead_stock_data), len(headers1))
    _autofit(ws1, {
        "SKU": 12, "Product Name": 36, "Collection": 22, "Category": 16,
        "Material": 12, "Warehouse": 18, "Qty On Hand": 14,
        "Days Since Last Sale": 20, "Stock Health Tier": 18,
        "Clearance Price": 16, "Unit Cost": 12, "Capital Tied Up": 16,
        "Recommendation": 38,
    })
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{len(dead_stock_data)+1}"

    # ── Sheet 2: Tuscany Collection Detail ─────────────────────────────
    ws2 = wb.create_sheet("Tuscany Collection Detail")
    headers2 = [
        "SKU", "Product Name", "Material", "Style Tags",
        "Unit Cost", "Clearance Price", "Competitor Avg Price", "Suggested Relaunch Price",
        "Qty On Hand", "Qty Reserved", "Available Stock",
        "Warehouse", "Received Date", "Days in Warehouse",
        "Potential Revenue (at Suggested Price)", "Margin at Suggested Price %",
    ]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, len(headers2))

    tuscany_detail = [
        ["SKU-0001", "Tuscany Raw Oak Dining Table", "raw oak",
         "organic-modern | warm-neutrals | raw-wood",
         320, 400, 1200, 899, 38, 0, 38,
         "WH-GA (Georgia)", "2024-07-30", 210,
         34162, "180.9%"],
        ["SKU-0002", "Tuscany Linen Dining Chair", "linen",
         "organic-modern | warm-neutrals | boucle",
         95, 120, 350, 249, 124, 0, 124,
         "WH-GA (Georgia)", "2024-07-30", 210,
         30876, "162.1%"],
    ]
    for r, row_data in enumerate(tuscany_detail, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c in (5, 6, 7, 8, 15):
                cell.number_format = '"$"#,##0'
            if c == 8:
                cell.font = GREEN_FONT
            if c == 15:
                cell.font = GREEN_FONT

    _style_data(ws2, len(tuscany_detail), len(headers2))
    _autofit(ws2, {
        "SKU": 12, "Product Name": 34, "Material": 12, "Style Tags": 40,
        "Unit Cost": 12, "Clearance Price": 16, "Competitor Avg Price": 20,
        "Suggested Relaunch Price": 22, "Qty On Hand": 14, "Qty Reserved": 14,
        "Available Stock": 16, "Warehouse": 18, "Received Date": 14,
        "Days in Warehouse": 18, "Potential Revenue (at Suggested Price)": 32,
        "Margin at Suggested Price %": 24,
    })
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Trend Match Analysis ──────────────────────────────────
    ws3 = wb.create_sheet("Trend Match Analysis")
    headers3 = [
        "Trending Style", "Style Tag", "Matching SKUs",
        "Top Matching Collection", "Match Score %", "Market Opportunity",
        "Action",
    ]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    _style_header(ws3, len(headers3))

    trend_match_data = [
        ["Organic Modern", "organic-modern", 47, "Tuscany Collection", 98,
         "🔥 HIGH — $1,200+ competitor price, massive consumer demand",
         "RELAUNCH as 'The Arden Collection' at $899"],
        ["Warm Neutrals", "warm-neutrals", 62, "Tuscany Collection", 95,
         "🔥 HIGH — aligns with top color palette trend",
         "Feature in The Arden Collection campaign"],
        ["Raw Wood", "raw-wood", 28, "Tuscany Collection", 97,
         "🔥 HIGH — raw oak = exact material match to trend",
         "Hero product in marketing visuals"],
        ["Boucle / Natural Textiles", "boucle", 19, "Organic Living", 85,
         "📈 MEDIUM-HIGH — linen chairs match boucle aesthetic",
         "Cross-promote with Tuscany chairs"],
        ["Scandinavian Minimal", "scandinavian", 34, "Nordic Frost", 72,
         "📊 MODERATE — stable demand, no spike",
         "Maintain current strategy"],
        ["Industrial Urban", "industrial", 29, "Metro Edge", 45,
         "📉 LOW — declining trend momentum",
         "Consider clearance pricing"],
        ["Coastal Relaxed", "coastal", 22, "Coastal Breeze", 58,
         "📊 MODERATE — seasonal, summer peak expected",
         "Plan seasonal promotion"],
    ]
    for r, row_data in enumerate(trend_match_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            if c == 5 and isinstance(val, (int, float)) and val >= 90:
                cell.font = GREEN_FONT
            # Highlight top matches
            if r <= 4:
                if c in (1, 4, 7):
                    cell.fill = ACCENT_FILL

    _style_data(ws3, len(trend_match_data), len(headers3))
    _autofit(ws3, {
        "Trending Style": 22, "Style Tag": 18, "Matching SKUs": 14,
        "Top Matching Collection": 24, "Match Score %": 14,
        "Market Opportunity": 52, "Action": 36,
    })
    ws3.freeze_panes = "A2"

    path = OUTPUT_DIR / "inventory_dead_stock_report.xlsx"
    wb.save(path)
    print(f"  ✓ {path}")
    return path


# =======================================================================
# Main
# =======================================================================
if __name__ == "__main__":
    print("Generating demo Excel files...")
    create_market_research()
    create_inventory_report()
    print("\nDone!")
