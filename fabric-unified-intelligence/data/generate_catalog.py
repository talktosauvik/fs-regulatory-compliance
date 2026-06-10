#!/usr/bin/env python3
"""
Generate a mock global product catalog (~500 SKUs) in both Excel (.xlsx)
and CSV (.csv) formats.

The first 50 SKUs match the BigQuery furniture_stock data exactly.
The remaining ~450 SKUs extend across 8 collections to simulate a
global retailer's full product universe.

Prerequisites:
  pip install openpyxl
"""

import csv
import random
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("ERROR: openpyxl is required.\n  pip install openpyxl")

random.seed(42)

OUTPUT_DIR = Path(__file__).parent
XLSX_PATH = OUTPUT_DIR / "global_product_catalog.xlsx"
CSV_PATH = OUTPUT_DIR / "global_product_catalog.csv"

# ---------------------------------------------------------------------------
# Reference data — shared with populate_tables.py
# ---------------------------------------------------------------------------
COLLECTIONS = {
    "Tuscany Collection": {
        "materials": ["raw oak", "linen"],
        "style_tags": ["organic-modern", "warm-neutrals", "raw-wood"],
        "color_palettes": ["Warm Honey / Natural Cream", "Golden Oak / Ivory Linen"],
    },
    "Organic Living": {
        "materials": ["bamboo", "linen", "teak"],
        "style_tags": ["organic-modern", "warm-neutrals", "boucle"],
        "color_palettes": ["Natural Bamboo / Soft White", "Teak Brown / Oat Linen"],
    },
    "Nordic Frost": {
        "materials": ["birch", "wool", "steel"],
        "style_tags": ["scandinavian", "minimalist", "cool-tones"],
        "color_palettes": ["Pale Birch / Arctic White", "Matte Steel / Frost Grey"],
    },
    "Metro Edge": {
        "materials": ["steel", "leather", "marble"],
        "style_tags": ["industrial", "urban", "minimalist"],
        "color_palettes": ["Black Steel / Charcoal Leather", "White Marble / Gunmetal"],
    },
    "Coastal Breeze": {
        "materials": ["rattan", "teak", "linen"],
        "style_tags": ["coastal", "relaxed", "warm-neutrals"],
        "color_palettes": ["Driftwood / Sandy Beige", "Ocean Teak / Washed Linen"],
    },
    "Heritage Craft": {
        "materials": ["walnut", "brass", "velvet"],
        "style_tags": ["traditional", "warm-tones", "handcrafted"],
        "color_palettes": ["Dark Walnut / Aged Brass", "Burgundy Velvet / Rich Walnut"],
    },
    "Urban Loft": {
        "materials": ["reclaimed wood", "iron", "concrete"],
        "style_tags": ["industrial", "rustic", "urban"],
        "color_palettes": ["Weathered Pine / Raw Iron", "Grey Concrete / Dark Iron"],
    },
    "Zen Garden": {
        "materials": ["bamboo", "stone", "cotton"],
        "style_tags": ["japanese", "minimalist", "natural"],
        "color_palettes": ["Light Bamboo / Pebble Grey", "Natural Cotton / Slate Stone"],
    },
}

CATEGORIES = [
    "Dining Table", "Chair", "Sofa", "Bookshelf", "Coffee Table",
    "Bed Frame", "Side Table", "Console", "Desk", "Nightstand",
]

# Dimension templates: (length_cm, width_cm, height_cm, weight_kg)
DIMENSION_TEMPLATES = {
    "Dining Table":  (200, 100, 75, 45.0),
    "Chair":         (55,  55,  90, 8.0),
    "Sofa":          (220, 95,  85, 55.0),
    "Bookshelf":     (90,  35,  180, 35.0),
    "Coffee Table":  (120, 60,  45, 20.0),
    "Bed Frame":     (210, 160, 40, 50.0),
    "Side Table":    (50,  50,  60, 10.0),
    "Console":       (130, 40,  80, 25.0),
    "Desk":          (140, 70,  75, 30.0),
    "Nightstand":    (50,  40,  55, 12.0),
}

DESCRIPTIONS = {
    "Dining Table": [
        "Generously proportioned dining table for family gatherings, crafted from {material}.",
        "Elegant {material} dining table with clean lines and a natural finish.",
        "Sturdy {material} table seating 6–8, ideal for modern dining rooms.",
    ],
    "Chair": [
        "Comfortable {material} dining chair with ergonomic lumbar support.",
        "Light yet durable {material} chair, perfect for everyday use.",
        "Sculptural {material} accent chair with a contemporary silhouette.",
    ],
    "Sofa": [
        "Deep-seated {material} sofa with plush cushions for ultimate relaxation.",
        "Modular {material} sofa that adapts to any living room layout.",
        "Classic 3-seater {material} sofa with solid hardwood frame.",
    ],
    "Bookshelf": [
        "Open-shelf {material} bookcase with five adjustable tiers.",
        "Tall {material} bookshelf with a ladder-style lean design.",
        "Wall-mounted {material} shelf system for books and décor.",
    ],
    "Coffee Table": [
        "Low-profile {material} coffee table with rounded edges.",
        "Round {material} coffee table combining form and function.",
        "Nesting {material} coffee table set for flexible living spaces.",
    ],
    "Bed Frame": [
        "Platform {material} bed frame with integrated headboard.",
        "Minimalist {material} bed frame with hidden under-bed storage.",
        "Queen-size {material} bed frame with slatted base support.",
    ],
    "Side Table": [
        "Compact {material} side table with a single drawer.",
        "Geometric {material} accent table for beside your sofa.",
        "Round {material} side table with a tapered pedestal base.",
    ],
    "Console": [
        "Slim {material} console table for entryways and hallways.",
        "Elegant {material} console with two open shelves below.",
        "Narrow {material} console table with brass-tipped legs.",
    ],
    "Desk": [
        "Spacious {material} writing desk with cable management.",
        "Standing-height adjustable {material} desk for home offices.",
        "Mid-century {material} desk with pencil drawer and hutch.",
    ],
    "Nightstand": [
        "Two-drawer {material} nightstand with soft-close hardware.",
        "Floating {material} nightstand with built-in wireless charging.",
        "Petite {material} nightstand perfect for small bedrooms.",
    ],
}

STATUSES = ["Active", "Active", "Active", "Active", "Discontinued", "Limited Edition"]

COLUMNS = [
    "sku", "product_name", "category", "collection", "material",
    "secondary_material", "color_palette", "style_tags", "msrp",
    "length_cm", "width_cm", "height_cm", "weight_kg",
    "description", "image_url", "status", "launch_year",
]


def _sku(n: int) -> str:
    return f"SKU-{n:04d}"


def _jitter(base: float, pct: float = 0.12) -> float:
    return round(base * random.uniform(1 - pct, 1 + pct), 1)


# ---------------------------------------------------------------------------
# Build catalog rows
# ---------------------------------------------------------------------------
rows: list[dict] = []

# ── 1. Hardcoded demo items (must match populate_tables.py exactly) ────
rows.append({
    "sku": "SKU-0001",
    "product_name": "Tuscany Raw Oak Dining Table",
    "category": "Dining Table",
    "collection": "Tuscany Collection",
    "material": "raw oak",
    "secondary_material": "",
    "color_palette": "Warm Honey / Natural Cream",
    "style_tags": "organic-modern | warm-neutrals | raw-wood",
    "msrp": 1099.00,
    "length_cm": 200, "width_cm": 100, "height_cm": 75, "weight_kg": 48.5,
    "description": "Handcrafted dining table in raw European oak with a natural matte finish. Seats 6–8 comfortably. Part of the Tuscany Collection inspired by Italian farmhouse living.",
    "image_url": "https://catalog.example.com/images/SKU-0001.jpg",
    "status": "Active",
    "launch_year": 2023,
})

rows.append({
    "sku": "SKU-0002",
    "product_name": "Tuscany Linen Dining Chair",
    "category": "Chair",
    "collection": "Tuscany Collection",
    "material": "linen",
    "secondary_material": "raw oak",
    "color_palette": "Ivory Linen / Golden Oak",
    "style_tags": "organic-modern | warm-neutrals | boucle",
    "msrp": 349.00,
    "length_cm": 55, "width_cm": 55, "height_cm": 92, "weight_kg": 7.8,
    "description": "Upholstered dining chair in natural linen with a raw oak frame. Ergonomic backrest and removable seat cushion. Part of the Tuscany Collection.",
    "image_url": "https://catalog.example.com/images/SKU-0002.jpg",
    "status": "Active",
    "launch_year": 2023,
})

# ── 2. BQ-matching items (SKU-0003 through SKU-0050) ──────────────────
# Reproduce the same logic as populate_tables.py background items
bq_collections = ["Organic Living", "Nordic Frost", "Metro Edge", "Coastal Breeze"]

for i in range(3, 51):
    coll_name = bq_collections[(i - 3) % len(bq_collections)]
    coll = COLLECTIONS[coll_name]
    cat = CATEGORIES[(i - 3) % len(CATEGORIES)]
    mat = coll["materials"][(i - 3) % len(coll["materials"])]
    dims = DIMENSION_TEMPLATES[cat]
    sec_mat = coll["materials"][(i - 2) % len(coll["materials"])] if len(coll["materials"]) > 1 else ""
    if sec_mat == mat:
        sec_mat = ""

    msrp = round(random.uniform(199, 2499), 2)
    desc = random.choice(DESCRIPTIONS[cat]).format(material=mat)

    rows.append({
        "sku": _sku(i),
        "product_name": f"{coll_name} {mat.title()} {cat}",
        "category": cat,
        "collection": coll_name,
        "material": mat,
        "secondary_material": sec_mat,
        "color_palette": random.choice(coll["color_palettes"]),
        "style_tags": " | ".join(random.sample(coll["style_tags"], k=min(2, len(coll["style_tags"])))),
        "msrp": msrp,
        "length_cm": _jitter(dims[0]),
        "width_cm": _jitter(dims[1]),
        "height_cm": _jitter(dims[2]),
        "weight_kg": _jitter(dims[3]),
        "description": desc,
        "image_url": f"https://catalog.example.com/images/{_sku(i)}.jpg",
        "status": random.choice(STATUSES),
        "launch_year": random.choice([2021, 2022, 2023, 2024, 2025]),
    })

# ── 3. Extended catalog items (SKU-0051 through ~SKU-0500) ────────────
sku_num = 51
all_collections = list(COLLECTIONS.keys())

VARIANT_LABELS = [
    "", "II", "Slim", "Wide", "Compact", "XL",
]

for coll_name in all_collections:
    coll = COLLECTIONS[coll_name]
    for cat in CATEGORIES:
        # 4–6 variants per collection × category → ~500 extended SKUs
        num_variants = random.randint(4, 6)
        for v in range(num_variants):
            mat = coll["materials"][v % len(coll["materials"])]
            dims = DIMENSION_TEMPLATES[cat]
            sec_idx = (v + 1) % len(coll["materials"])
            sec_mat = coll["materials"][sec_idx] if coll["materials"][sec_idx] != mat else ""

            msrp = round(random.uniform(149, 2999), 2)
            desc = random.choice(DESCRIPTIONS[cat]).format(material=mat)
            variant_label = f" {VARIANT_LABELS[v % len(VARIANT_LABELS)]}" if v > 0 else ""

            rows.append({
                "sku": _sku(sku_num),
                "product_name": f"{coll_name} {mat.title()} {cat}{variant_label}",
                "category": cat,
                "collection": coll_name,
                "material": mat,
                "secondary_material": sec_mat,
                "color_palette": coll["color_palettes"][v % len(coll["color_palettes"])],
                "style_tags": " | ".join(coll["style_tags"]),
                "msrp": msrp,
                "length_cm": _jitter(dims[0]),
                "width_cm": _jitter(dims[1]),
                "height_cm": _jitter(dims[2]),
                "weight_kg": _jitter(dims[3]),
                "description": desc,
                "image_url": f"https://catalog.example.com/images/{_sku(sku_num)}.jpg",
                "status": random.choice(STATUSES),
                "launch_year": random.choice([2020, 2021, 2022, 2023, 2024, 2025]),
            })
            sku_num += 1

print(f"  Total catalog rows: {len(rows)}")

# ---------------------------------------------------------------------------
# Write CSV
# ---------------------------------------------------------------------------
with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=COLUMNS)
    writer.writeheader()
    writer.writerows(rows)
print(f"  ✓ CSV written to {CSV_PATH}")

# ---------------------------------------------------------------------------
# Write Excel with formatting
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Product Catalog"

# Header styling
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)

# Write headers
for col_idx, col_name in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Write data
data_font = Font(name="Calibri", size=10)
alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

for row_idx, row_data in enumerate(rows, start=2):
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
        cell.font = data_font
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill

# Auto-fit column widths (approximate)
COL_WIDTHS = {
    "sku": 12, "product_name": 38, "category": 16, "collection": 22,
    "material": 18, "secondary_material": 20, "color_palette": 30,
    "style_tags": 40, "msrp": 10, "length_cm": 11, "width_cm": 10,
    "height_cm": 11, "weight_kg": 11, "description": 65,
    "image_url": 45, "status": 16, "launch_year": 12,
}
for col_idx, col_name in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

wb.save(XLSX_PATH)
print(f"  ✓ Excel written to {XLSX_PATH}")
print("\nDone!")
